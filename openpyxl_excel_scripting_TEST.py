#!/usr/bin/python

from openpyxl import Workbook

def get_using_cells(worksheet):
    ws = worksheet
    customer_using = {}
    for cell in ws['A'][3:]:
        customer_using[cell.value] = []
        for flag in ws[cell.row]:
            if flag.value == 'Using':
                customer_using[cell.value].append(flag.coordinate)

    return customer_using

def get_cust_count_table(worksheet):
    count_table = {}
    for cell in worksheet['A'][2:]:
        count_table[cell.value] = []
        for n in worksheet[cell.row]
            if 



def main(filename):
    wb = openpyxl.load_workbook(filename)
    customer_using = get_using_cells(wb.worksheets[1])

    for customer in customer_using.keys():


    using_cells
    for A in ws2_carrier_cells:
        carrier_name = A.value
        cur_row_num = A.row

        for B in ws0_carrier_cells:
            relevant_col_ids = []
            if B.value == carrier_name:
               for C in wb.worksheets[0][cell.row]:
                   if C.value == 'Using':
                       relevant_columns.append(C.col_idx)
               for col in relevant_col_ids:
                    for car_cell in ws1_carrier_cells:
                        if car_cell.value == carrier_name:


                        




if __name__ == "__main__":
    filename = sys.argv[1]
    main(filename)